#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
import os


def combine():
    wb = load_workbook('courses.xlsx')
    ws_name = wb.sheetnames
    sheet_1 = wb[ws_name[0]]
    sheet_2 = wb[ws_name[1]]
    sheet_3 = wb.create_sheet(title='combine')
    row_list = []
    for x in range(len(sheet_1['B'])):
        for y in range(len(sheet_2['B'])):
            if sheet_1['B'][x].value == sheet_2['B'][y].value:
                row_list = [sheet_1['A'][x].value, sheet_1['B'][x].value,
                            sheet_1['C'][x].value, sheet_2['C'][x].value]
                sheet_3.append(row_list)
    wb.save('courses.xlsx')


def split():
    year_list = []
    sheets_title = []
    file_list = []
    row_list = []
    wb_old = load_workbook('courses.xlsx')
    ws_combine = wb_old['combine']
    for x in range(2, len(ws_combine['A'])):
        if ws_combine['A'][x].value.year in year_list:
            continue
        else:
            year_list.append(ws_combine['A'][x].value.year)
    for x in range(len(year_list)):
        sheets_title.append(str(year_list[x]))
        file_list.append(str(year_list[x]) + '.xlsx')
    for x in range(2, len(ws_combine['A'])):
        i = year_list.index(ws_combine['A'][x].value.year)
        if os.path.isfile(os.path.dirname('courses.xls') + file_list[i]):
            wb = load_workbook(file_list[i])
            ws = wb[sheets_title[i]]
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheets_title[i]
        row_list = [ws_combine['A'][x].value, ws_combine['B'][x].value,
                    ws_combine['C'][x].value, ws_combine['D'][x].value]
        ws.append(row_list)
        wb.save(file_list[i])


if __name__ == '__main__':
    combine()
    split()
